import tkinter as tk
from tkinter import ttk, messagebox, Toplevel
from openpyxl import load_workbook

# Cargar el archivo Excel
archivo_excel = "tiempos_trabajo.xlsx"

class App:
    def __init__(self, root):
        self.root = root
        self.root.title("Gestión de tiempos de trabajo")
        
        # Crear campos de entrada
        self.create_widgets()

        # Mostrar datos actuales de la hoja
        self.load_data()

    def create_widgets(self):
        # Crear etiquetas y entradas para los campos
        fields = ['TICKETS', 'TAREA', 'ORGANISMO', 'TIPO', 'TIEMPO', 'NOTAS']
        self.entries = {}
        
        for idx, field in enumerate(fields):
            label = tk.Label(self.root, text=field)
            label.grid(row=idx, column=0, padx=5, pady=5)
            entry = tk.Entry(self.root)
            entry.grid(row=idx, column=1, padx=5, pady=5)
            self.entries[field] = entry

        # Botón para insertar datos
        btn_insert = tk.Button(self.root, text="Insertar", command=self.insert_data)
        btn_insert.grid(row=len(fields), column=0, padx=5, pady=5)
        
        # Botón para cargar datos
        btn_load = tk.Button(self.root, text="Cargar Datos", command=self.load_data)
        btn_load.grid(row=len(fields), column=1, padx=5, pady=5)
        
        # Botón para borrar un registro
        btn_delete = tk.Button(self.root, text="Borrar Selección", command=self.delete_data)
        btn_delete.grid(row=len(fields), column=2, padx=5, pady=5)

        # Crear la tabla para mostrar los datos
        self.tree = ttk.Treeview(self.root, columns=fields, show='headings')
        for field in fields:
            self.tree.heading(field, text=field)
            self.tree.column(field, anchor='center')
        self.tree.grid(row=len(fields)+1, column=0, columnspan=3, padx=5, pady=5)
        
        # Añadir scrollbar
        scrollbar = ttk.Scrollbar(self.root, orient='vertical', command=self.tree.yview)
        self.tree.configure(yscroll=scrollbar.set)
        scrollbar.grid(row=len(fields)+1, column=3, sticky='ns')

        # Etiqueta para mostrar el total de tiempo
        self.total_time_label = tk.Label(self.root, text="Total Tiempo: 0 horas")
        self.total_time_label.grid(row=len(fields)+2, column=0, columnspan=3, padx=5, pady=5)

        # Doble clic para editar
        self.tree.bind("<Double-1>", self.edit_data)

    def load_data(self):
        # Cargar el archivo de Excel y mostrar los datos
        try:
            workbook = load_workbook(archivo_excel)
            sheet = workbook.active
            self.tree.delete(*self.tree.get_children())  # Limpiar la tabla
            total_time = 0  # Inicializar el total de tiempos
            for row in sheet.iter_rows(min_row=4, values_only=True):
                if any(row):  # Solo añadir filas que tengan datos
                    self.tree.insert('', 'end', values=row)
                    try:
                        total_time += float(row[4])  # Sumar el valor de "TIEMPO" (columna F3)
                    except (ValueError, TypeError):
                        pass  # Ignorar filas que no tengan un valor válido en "TIEMPO"
            
            # Mostrar el total de tiempo en la interfaz
            self.total_time_label.config(text=f"Total Tiempo: {total_time:.2f} horas")

            # Guardar los cambios en el archivo Excel
            workbook.save(archivo_excel)
            workbook.close()

        except FileNotFoundError:
            messagebox.showerror("Error", f"No se encontró el archivo {archivo_excel}")
        except Exception as e:
            messagebox.showerror("Error", f"Ocurrió un error al cargar los datos: {str(e)}")

    def insert_data(self):
        # Insertar nuevos datos en la hoja de Excel
        try:
            workbook = load_workbook(archivo_excel)
            sheet = workbook.active
            nueva_fila = [self.entries[field].get() for field in self.entries]
            
            if not all(nueva_fila[:5]):  # Los primeros 5 campos son obligatorios
                messagebox.showerror("Error", "Todos los campos (excepto Notas) deben ser completados.")
                return
            
            # Insertar en la primera fila vacía a partir de la fila 4
            sheet.append(nueva_fila)
            workbook.save(archivo_excel)
            workbook.close()
            self.load_data()  # Recargar los datos en la tabla
        except FileNotFoundError:
            messagebox.showerror("Error", f"No se encontró el archivo {archivo_excel}")
        except Exception as e:
            messagebox.showerror("Error", f"Ocurrió un error al insertar los datos: {str(e)}")

    def delete_data(self):
        # Borrar el registro seleccionado
        selected_item = self.tree.selection()
        if selected_item:
            values = self.tree.item(selected_item)['values']
            confirm = messagebox.askyesno("Confirmación", f"¿Seguro que quieres eliminar el registro {values}?")
            if confirm:
                try:
                    workbook = load_workbook(archivo_excel)
                    sheet = workbook.active
                    for idx, row in enumerate(sheet.iter_rows(min_row=4, values_only=True), start=4):
                        if list(row) == list(values):
                            sheet.delete_rows(idx, 1)
                            break
                    workbook.save(archivo_excel)
                    workbook.close()
                    self.load_data()  # Recargar los datos
                except FileNotFoundError:
                    messagebox.showerror("Error", f"No se encontró el archivo {archivo_excel}")
                except Exception as e:
                    messagebox.showerror("Error", f"Ocurrió un error al borrar el registro: {str(e)}")
        else:
            messagebox.showerror("Error", "Debes seleccionar un registro para eliminar.")

    def edit_data(self, event):
        # Obtener el elemento seleccionado
        selected_item = self.tree.selection()
        if not selected_item:
            messagebox.showerror("Error", "Debes seleccionar un registro para editar.")
            return

        # Obtener los valores seleccionados
        values = self.tree.item(selected_item)['values']

        # Crear una ventana emergente para editar
        edit_window = Toplevel(self.root)
        edit_window.title("Editar Registro")
        
        fields = ['TICKETS', 'TAREA', 'ORGANISMO', 'TIPO', 'TIEMPO', 'NOTAS']
        new_entries = {}
        
        for idx, field in enumerate(fields):
            label = tk.Label(edit_window, text=field)
            label.grid(row=idx, column=0, padx=5, pady=5)
            entry = tk.Entry(edit_window)
            entry.grid(row=idx, column=1, padx=5, pady=5)
            entry.insert(0, values[idx])  # Llenar con el valor actual
            new_entries[field] = entry

        # Botón para guardar los cambios
        btn_save = tk.Button(edit_window, text="Guardar", 
                             command=lambda: self.save_edited_data(new_entries, values, edit_window))
        btn_save.grid(row=len(fields), column=0, columnspan=2, pady=10)

    def save_edited_data(self, new_entries, old_values, edit_window):
        try:
            # Cargar el archivo Excel
            workbook = load_workbook(archivo_excel)
            sheet = workbook.active
            
            # Convertir los valores nuevos en una lista
            edited_values = [new_entries[field].get() for field in new_entries]
            
            # Buscar y reemplazar la fila antigua con la fila nueva
            for idx, row in enumerate(sheet.iter_rows(min_row=4, values_only=True), start=4):
                if list(row) == list(old_values):
                    for col_idx, value in enumerate(edited_values, start=1):
                        sheet.cell(row=idx, column=col_idx, value=value)
                    break
            
            workbook.save(archivo_excel)
            workbook.close()

            # Recargar los datos en la tabla
            self.load_data()

            # Cerrar la ventana de edición
            edit_window.destroy()

        except FileNotFoundError:
            messagebox.showerror("Error", f"No se encontró el archivo {archivo_excel}")
        except Exception as e:
            messagebox.showerror("Error", f"Ocurrió un error al guardar los cambios: {str(e)}")

if __name__ == "__main__":
    root = tk.Tk()
    app = App(root)
    root.mainloop()
